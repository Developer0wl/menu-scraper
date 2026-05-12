"""
Import allergen data from Perplexity Excel for all-CNV chains.
Writes checkpoint JSON files compatible with the Allerva scraper format.
"""
import json
import openpyxl
from datetime import datetime, timezone
from pathlib import Path

EXCEL_PATH = Path(__file__).parent / "Allerva_Chain_Database_Complete_1 (1).xlsx"
CHECKPOINTS_DIR = Path(__file__).parent / "checkpoints"

NOW = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")

# Map our checkpoint key → Perplexity sheet name
CNV_CHAINS = {
    "wafflehouse":   "Waffle House",
    "einsteinbros":  "Einstein Bros Bagels",
    "tacobell":      "Taco Bell",
    "redrobin":      "Red Robin",
    "pandaexpress":  "Panda Express",
    "potbelly":      "Potbelly Sandwich Works",
    "firehousesubs": "Firehouse Subs",
    "smashburger":   "Smashburger",
    "raisingcanes":  "Raising Cane's",
    "goldencorral":  "Golden Corral",
    "jerseymikes":   "Jersey Mike's Subs",
}

# 24 chains currently in the app but missing from our scraper
APP_CHAINS = {
    "burgerking":        "Burger King",
    "wendys":            "Wendy's",
    "kfc":               "KFC",
    "popeyes":           "Popeyes",
    "sonic":             "Sonic",
    "arbys":             "Arby's",
    "jackinthebox":      "Jack in the Box",
    "dairyqueen":        "Dairy Queen",
    "shakeshack":        "Shake Shack",
    "dominos":           "Domino's",
    "papajohns":         "Papa John's",
    "pizzahut":          "Pizza Hut",
    "olivegarden":       "Olive Garden",
    "starbucks":         "Starbucks",
    "dunkin":            "Dunkin",
    "panerabread":       "Panera Bread",
    "applebees":         "Applebee's",
    "buffalowildwings":  "Buffalo Wild Wings",
    "chilis":            "Chili's",
    "dennys":            "Denny's",
    "ihop":              "IHOP",
    "outbacksteakhouse": "Outback Steakhouse",
    "redlobster":        "Red Lobster",
    "cheesecakefactory": "The Cheesecake Factory",
}

ALLERGEN_COLS = {
    "milk":      4,   # col E (0-indexed from col B at idx 1)
    "eggs":      5,
    "fish":      6,
    "shellfish": 7,
    "treeNuts":  8,
    "peanuts":   9,
    "wheat":     10,
    "soy":       11,
    "sesame":    12,
}
# col B=1, C=2, D=3, E=4 ... O=14  (0-indexed from the full row tuple)


def parse_bool(val):
    if val is None:
        return "COULD_NOT_VERIFY"
    v = str(val).strip().upper()
    if v == "TRUE":
        return "TRUE"
    if v == "FALSE":
        return "FALSE"
    return "COULD_NOT_VERIFY"


def parse_cross(val):
    if val is None:
        return "COULD_NOT_VERIFY"
    v = str(val).strip().upper()
    if "YES" in v or "TRUE" in v:
        return "TRUE"
    if "NO" in v or "FALSE" in v:
        return "FALSE"
    return "COULD_NOT_VERIFY"


def extract_sheet(ws, chain_key, sheet_name):
    rows_data = []
    header_found = False
    data_start = None

    all_rows = list(ws.iter_rows(values_only=True))

    # Find header row (contains "Item Name")
    for i, row in enumerate(all_rows):
        if any(str(v).strip() == "Item Name" for v in row if v is not None):
            header_found = True
            data_start = i + 1
            break

    if not header_found:
        print(f"  WARNING: No header row found in '{sheet_name}'")
        return []

    source_url = None
    row_num = 0

    for row in all_rows[data_start:]:
        # Skip empty rows or footer rows
        if row[1] is None:
            continue
        # Stop if col B is not a number (footer)
        try:
            int(row[1])
        except (TypeError, ValueError):
            continue

        item_name = str(row[3]).strip() if row[3] is not None else ""
        if not item_name:
            continue

        menu_category = str(row[2]).strip() if row[2] is not None else "Menu"
        src_url = str(row[14]).strip() if len(row) > 14 and row[14] is not None else ""
        if src_url and not src_url.startswith("http"):
            src_url = ""
        if src_url:
            source_url = src_url

        row_num += 1
        entry = {
            "rowNum": row_num,
            "menuCategory": menu_category,
            "itemName": item_name,
            "crossContact": parse_cross(row[13] if len(row) > 13 else None),
            "sourceUrl": src_url or f"https://www.{chain_key}.com",
            "scrapeDate": NOW,
            "confidence": "HIGH",
            "sourceText": f"Imported from SafeBite/Perplexity allergen database (March 2026). Source: {src_url}",
            "milk":      parse_bool(row[4]),
            "eggs":      parse_bool(row[5]),
            "fish":      parse_bool(row[6]),
            "shellfish": parse_bool(row[7]),
            "treeNuts":  parse_bool(row[8]),
            "peanuts":   parse_bool(row[9]),
            "wheat":     parse_bool(row[10]),
            "soy":       parse_bool(row[11]),
            "sesame":    parse_bool(row[12]),
        }
        rows_data.append(entry)

    return rows_data


def main():
    import sys
    mode = sys.argv[1] if len(sys.argv) > 1 else "app"
    chains_to_import = APP_CHAINS if mode == "app" else CNV_CHAINS
    print(f"Loading {EXCEL_PATH.name}... (mode={mode}, {len(chains_to_import)} chains)")

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    available_sheets = set(wb.sheetnames)

    results = {}
    for chain_key, sheet_name in chains_to_import.items():
        if sheet_name not in available_sheets:
            print(f"  SKIP {chain_key}: sheet '{sheet_name}' not found")
            continue

        ws = wb[sheet_name]
        rows = extract_sheet(ws, chain_key, sheet_name)

        if not rows:
            print(f"  SKIP {chain_key}: 0 rows extracted from '{sheet_name}'")
            continue

        true_count = sum(
            1 for r in rows
            if any(r[a] == "TRUE" for a in ["milk","eggs","fish","shellfish","treeNuts","peanuts","wheat","soy","sesame"])
        )

        checkpoint = {
            "chainName": chain_key,
            "savedAt": NOW,
            "rowCount": len(rows),
            "rows": rows,
        }

        out_path = CHECKPOINTS_DIR / f"{chain_key}.json"
        out_path.write_text(json.dumps(checkpoint, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"  OK {chain_key}: {len(rows)} rows, {true_count} with TRUE allergens -> {out_path.name}")
        results[chain_key] = {"rows": len(rows), "true": true_count}

    print(f"\nDone. {len(results)}/{len(CNV_CHAINS)} chains imported.")
    for k, v in results.items():
        print(f"  {k}: {v['rows']} rows, {v['true']} TRUE")


if __name__ == "__main__":
    main()
